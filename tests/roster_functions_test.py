# import pytest
import os
import tempfile
import csv
import openpyxl

from src.process_excel_file import check_sheet_titles
from src.process_excel_file import separate_first_last_name
from src.process_excel_file import process_teams
from src.process_excel_file import write_rosters_to_csv
from src.process_excel_file import fix_team_name

# Test for checking the sheet titles
def test_check_sheet_titles():
    # Create a test Excel file
    wb = openpyxl.Workbook()

    # Add a sheet with all required headers
    sheet1 = wb.create_sheet(title='Sheet1')
    sheet1.append(['Firstname', 'Lastname', 'Team', 'Sweater'])
    assert check_sheet_titles(sheet1) == True

    # Add a sheet with all required headers in lowercase
    sheet2 = wb.create_sheet(title='Sheet2')
    sheet2.append(['firstname', 'lastname', 'team', 'sweater'])
    assert check_sheet_titles(sheet2) == True

    # Add a sheet with all required headers in alternative variations
    sheet3 = wb.create_sheet(title='Sheet3')
    sheet3.append(['First Name', 'Last Name', 'Team', 'Sweater'])
    assert check_sheet_titles(sheet3) == True

    # Add a sheet with missing headers
    sheet4 = wb.create_sheet(title='Sheet4')
    sheet4.append(['Firstname', 'Lastname', 'Team'])
    assert check_sheet_titles(sheet4) == False

    # Test for three columns
    sheet5 = wb.create_sheet(title='Sheet5')
    sheet5.append(['Name', 'Team', 'Sweater'])
    assert check_sheet_titles(sheet5) == True

    # Test for a column with NoneType
    sheet6 = wb.create_sheet(title='Sheet6')
    sheet6.append(['Name', 'Team', 'Sweater', None])
    assert check_sheet_titles(sheet6) == True

    # Test for a column with an empty string
    sheet7 = wb.create_sheet(title='Sheet7')
    sheet7.append(['Name', 'Team', 'Sweater', ''])
    assert check_sheet_titles(sheet7) == False

def test_fix_team_name():
    assert fix_team_name('11u AAA') == '11U AAA'
    assert fix_team_name('12u red') == '12U Red'
    assert fix_team_name('13Uaaa') == '13U AAA'
    assert fix_team_name('16U green giants') == '16U Green Giants'
    assert fix_team_name('U9 orange panthers') == 'U9 Orange Panthers'
    assert fix_team_name('14U BLACK') == '14U Black'
    assert fix_team_name('10U Blue') == '10U Blue'
    assert fix_team_name('11URED') == '11U Red'


def test_separate_first_last_name():
    assert separate_first_last_name("John Smith") == ("John", "Smith")
    assert separate_first_last_name("John") == ("John", "")
    assert separate_first_last_name("John Van Damme") == ("John", "Van Damme")
    assert separate_first_last_name(
        "John (Jack) Smith") == ("John (Jack)", "Smith")
    assert separate_first_last_name("John Smith Jr.") == ("John", "Smith Jr.")
    assert separate_first_last_name("John Smith III") == ("John", "Smith III")
    # Test for three names with a middle initial
    assert separate_first_last_name('John Q Doe') == ('John Q', 'Doe')


def test_process_teams():
    # Setup: create a mock workbook and sheet
    # Create a test Excel file
    workbook = openpyxl.Workbook()

    # Add a sheet with all required headers
    sheet_name = workbook.create_sheet(title='Sheet1')
    sheet_name.append(['Fullname', 'Team', 'Sweater'])

    # Add player data to the sheet
    sheet_name.append(['John Smith III', 'Team A', '12'])
    sheet_name.append(['Bob Johnson', 'Team A', '67'])

    # Add blank lines to the sheet
    sheet_name.append(['', '', ''])
    sheet_name.append(['Jane Van Doe', 'Team B', '4'])


    # Call the function under test
    rosters = process_teams(workbook, 'Sheet1')

    # Assert that the function returns the expected value
    assert rosters == {
        'Team A': [
            {'Firstname': 'John', 'Lastname': 'Smith III',
                'Team': 'Team A', 'Sweater': '12'},
            {'Firstname': 'Bob', 'Lastname': 'Johnson',
                'Team': 'Team A', 'Sweater': '67'},
        ],
        'Team B': [
            {'Firstname': 'Jane', 'Lastname': 'Van Doe',
                'Team': 'Team B', 'Sweater': '4'},
        ]
    }

# test for Firstname, Lastname, Team, Sweater already in the sheet
def test_process_teams_first_last_name():
    # Setup: create a mock workbook and sheet
    # Create a test Excel file
    workbook = openpyxl.Workbook()

    # Add a sheet with all required headers
    sheet_name = workbook.create_sheet(title='Sheet1')
    sheet_name.append(['Firstname', 'Lastname', 'Team', 'Sweater'])

    # Add player data to the sheet
    sheet_name.append(['John', 'Smith III', 'Team A', '12'])
    sheet_name.append(['Bob', 'Johnson', 'Team A', '67'])

    # Add blank lines to the sheet
    sheet_name.append(['', '', '', ''])
    sheet_name.append(['Jane', 'Van Doe', 'Team B', '4'])


    # Call the function under test
    rosters = process_teams(workbook, 'Sheet1')

    # Assert that the function returns the expected value
    assert rosters == {
        'Team A': [
            {'Firstname': 'John', 'Lastname': 'Smith III',
                'Team': 'Team A', 'Sweater': '12'},
            {'Firstname': 'Bob', 'Lastname': 'Johnson',
                'Team': 'Team A', 'Sweater': '67'},
        ],
        'Team B': [
            {'Firstname': 'Jane', 'Lastname': 'Van Doe',
                'Team': 'Team B', 'Sweater': '4'},
        ]
    }

# Test for workbook having multiple sheets
def test_process_teams_multiple_sheets():
    # Setup: create a mock workbook and sheet
    # Create a test Excel file
    workbook = openpyxl.Workbook()

    # Add a sheet with all required headers
    sheet_name = workbook.create_sheet(title='Sheet1')
    sheet_name.append(['Fullname', 'Team', 'Sweater'])

    # Add player data to the sheet
    sheet_name.append(['John Smith III', 'Team A', '12'])
    sheet_name.append(['Bob Johnson', 'Team A', '67'])

    # Add blank lines to the sheet
    sheet_name.append(['', '', ''])
    sheet_name.append(['Jane Van Doe', 'Team B', '4'])

    # Add a second sheet with all required headers
    sheet_name2 = workbook.create_sheet(title='Sheet2')
    sheet_name2.append(['Fullname', 'Team', 'Sweater'])

    # Add player data to the sheet
    sheet_name2.append(['John Smith III', 'Team A', '12'])
    sheet_name2.append(['Bob Johnson', 'Team A', '67'])

    # Add blank lines to the sheet
    sheet_name2.append(['', '', ''])
    sheet_name2.append(['Jane Van Doe', 'Team B', '4'])

    # Add a third sheet with all required headers
    sheet_name3 = workbook.create_sheet(title='Sheet3')
    sheet_name3.append(['Fullname', 'Team', 'Sweater'])

    # Add blank line and player data to the sheet
    sheet_name3.append(['', '', ''])
    sheet_name3.append(['John Smith III', 'Team C', '12'])

    teams = {}
    # Call the function under test
    for sheet in workbook:
        roster = process_teams(workbook, sheet.title)
        # Add roster to a list of rosters
        teams.update(roster)
    

    # Assert that the function returns the expected value
    assert teams == {
        'Team A': [
            {'Firstname': 'John', 'Lastname': 'Smith III',
                'Team': 'Team A', 'Sweater': '12'},
            {'Firstname': 'Bob', 'Lastname': 'Johnson',
                'Team': 'Team A', 'Sweater': '67'},
        ],
        'Team B': [
            {'Firstname': 'Jane', 'Lastname': 'Van Doe',
                'Team': 'Team B', 'Sweater': '4'},
        ],
        'Team C': [
            {'Firstname': 'John', 'Lastname': 'Smith III',
                'Team': 'Team C', 'Sweater': '12'}
        ]
    }

# Test write_rosters_to_csv function
def test_write_rosters_to_csv():
    # Create a temporary directory
    with tempfile.TemporaryDirectory() as tmpdir:
        # Set the current working directory to the temporary directory
        os.chdir(tmpdir)

        # Create some test data
        rosters = {
            'Team 1': [
                {'Firstname': 'John', 'Lastname': 'Doe',
                    'Team': 'Team 1', 'Sweater': '1'},
                {'Firstname': 'Jane', 'Lastname': 'Doe',
                    'Team': 'Team 1', 'Sweater': '2'},
            ],
            'Team 2': [
                {'Firstname': 'Bob', 'Lastname': 'Smith',
                    'Team': 'Team 2', 'Sweater': '3'},
                {'Firstname': 'Alice', 'Lastname': 'Smith',
                    'Team': 'Team 2', 'Sweater': '4'},
            ]
        }

        # Call the function under test
        write_rosters_to_csv(rosters)

        # Verify that the CSV files were created
        assert os.path.exists('csv/Team_1.csv')
        assert os.path.exists('csv/Team_2.csv')

        # Verify the contents of the CSV files
        with open('csv/Team_1.csv', 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            assert rows == [
                {'Firstname': 'John', 'Lastname': 'Doe',
                    'Team': 'Team 1', 'Sweater': '1'},
                {'Firstname': 'Jane', 'Lastname': 'Doe',
                    'Team': 'Team 1', 'Sweater': '2'},
            ]

        with open('csv/Team_2.csv', 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            assert rows == [
                {'Firstname': 'Bob', 'Lastname': 'Smith',
                    'Team': 'Team 2', 'Sweater': '3'},
                {'Firstname': 'Alice', 'Lastname': 'Smith',
                    'Team': 'Team 2', 'Sweater': '4'},
            ]
        # Change back to the original directory
        os.chdir('..')
    # End of temporary directory

