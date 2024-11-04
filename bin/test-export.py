#!/usr/bin/env python3.9
# -*- coding: utf-8 -*-
#
# File: export_csv_from_excel.py
# Author: Scott Stonefield
#
# Read in an Excel file and export each sheet as a CSV file
#
import csv
import openpyxl
import os
import argparse
import sys

# Initial player count
player_count = 0

# Header constants needs to match the headers in the Excel file
HEADER_FIRSTNAME = 'Players First Name'
HEADER_LASTNAME = 'Players Last Name'
HEADER_TEAM = 'Players Team'
HEADER_SWEATER = 'Jersey Number'

# Create mapping for header names to original header names
header_mapping = {
    HEADER_FIRSTNAME: 'Firstname',
    HEADER_LASTNAME: 'Lastname',
    HEADER_TEAM: 'Team',
    HEADER_SWEATER: 'Sweater'
}

def clean_row(row, headers):
    global player_count

    firstname_index = headers.index(HEADER_FIRSTNAME)
    lastname_index = headers.index(HEADER_LASTNAME)
    sweater_index = headers.index(HEADER_SWEATER)

    row = list(row)

    if row[firstname_index] is None:
        print(f"Error: {HEADER_FIRSTNAME} is empty in row '{row[lastname_index]}' -- team: {row[headers.index(HEADER_TEAM)]}")
    else:
        row[firstname_index] = row[firstname_index].strip()
        # Make sure the first letter of the first name is upper case and the rest are lower case
        row[firstname_index] = row[firstname_index].lower()
        row[firstname_index] = row[firstname_index][0].upper() + row[firstname_index][1:]

    if row[lastname_index] is None:
        print(f"Error: {HEADER_LASTNAME} is empty in row '{row[firstname_index]}' -- team: {row[headers.index(HEADER_TEAM)]}")
    else:
        row[lastname_index] = row[lastname_index].strip()
        # Make sure the first letter of the last name is upper case and the rest are lower case
        row[lastname_index] = row[lastname_index].lower()
        row[lastname_index] = row[lastname_index][0].upper() + row[lastname_index][1:]

    # Check and clean Sweater number
    if row[sweater_index] is not None:
        row[sweater_index] = ''.join(filter(str.isdigit, str(row[sweater_index])))
        if not row[sweater_index]:
            print(f"Error: Invalid {HEADER_SWEATER} number in row '{row[firstname_index]} {row[lastname_index]}' -- team: {row[headers.index(HEADER_TEAM)]}")
            row[sweater_index] = '00'
    else:
        print(f"Error: {HEADER_SWEATER} number is empty in row '{row[firstname_index]} {row[lastname_index]}' -- team: {row[headers.index(HEADER_TEAM)]}")
        row[sweater_index] = '00'

    player_count += 1
    return row

def export_by_workbook(filename):
    global player_count
    
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)
    
    player_count = 0
    # Iterate through each sheet in the workbook
    for sheet in wb:
        # Get the first row (which should contain the headers)
        headers = [cell.value for cell in next(sheet.rows)]

        # Check if the required headers are present
        if not all(header in headers for header in header_mapping.keys()):
            print(f"Error: Required headers not found in sheet '{sheet.title}'")
            continue
        
        # If sheet.title has a space in the name, replace it with an underscore
        sheet.title = sheet.title.replace(' ', '_')
        
        # replace any characters that are not allowed in a filename with an underscore
        sheet.title = ''.join(c if c.isalnum() else '_' for c in sheet.title)
        
        # Create a CSV file for the sheet
        with open(f'csv/{sheet.title}.csv', 'w', newline='') as csv_file:
            # Create a CSV writer
            writer = csv.writer(csv_file)
            # Write the mapped headers to the CSV file
            writer.writerow([header_mapping[header] for header in headers])
            # Iterate through the rest of the rows in the sheet and write them to the CSV file
            for row in sheet.iter_rows(min_row=2, values_only=True):  # start from the second row to skip the header
                cleaned_row = clean_row(row, headers)
                writer.writerow(cleaned_row)



def get_headers(sheet):
    return [cell.value for cell in next(sheet.rows)]
    
def validate_headers(headers):
    return all(header in headers for header in header_mapping.keys())

def sanitize_filename(name):
    if name is not None:
        name = name.replace(' ', '_')
        name = ''.join(c if c.isalnum() else '_' for c in name)
    return name

def open_csv_writer(team_name):
    sanitized_name = sanitize_filename(team_name)
    csv_file = open(f'csv/{sanitized_name}.csv', 'w', newline='')
    writer = csv.writer(csv_file)
    writer.writerow([header_mapping[header] for header in headers])
    return csv_file, writer

def process_rows(sheet, headers):
    team_index = headers.index(HEADER_TEAM)
    last_team_name = None
    csv_file = None
    writer = None
    team_player_count = {}

def export_by_team(filename):
    global player_count
        
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[team_index] is None:
            continue

        team_name = row[team_index]
        if team_name != last_team_name:
            if csv_file is not None:
                csv_file.close()
            csv_file, writer = open_csv_writer(team_name)
            print(f"Team:", team_name)
            last_team_name = team_name
            team_player_count[team_name] = 0

        cleaned_row = clean_row(row, headers)
        writer.writerow(cleaned_row)
        team_player_count[team_name] += 1
    
    if csv_file is not None:
        csv_file.close()

    return team_player_count
    
    wb = openpyxl.load_workbook(filename)
    player_count = 0
    flagged_teams = []

    for sheet in wb:
        headers = get_headers(sheet)
        if not validate_headers(headers):
            print(f"Error: Required headers not found in sheet '{sheet.title}'")
            continue
        team_player_count = process_rows(sheet, headers)
        flagged_teams.extend([team for team, count in team_player_count.items() if count <= 5])

    if flagged_teams:
        print(f"\nFlagged Teams (5 or fewer players): {', '.join(flagged_teams)}")
    else:
        print("\nNo teams with 5 or fewer players were found.")


def main():
    global player_count
    
    # Command line argument parsing
    parser = argparse.ArgumentParser(description='Export Excel file to CSV.')
    parser.add_argument('filename', help='Name of the Excel file to export.')
    parser.add_argument('--by-team', action='store_true', help='If set, export by team. Otherwise, export by workbook.')

    args = parser.parse_args()

    # Create the "csv" directory if it doesn't exist
    if not os.path.exists("csv"):
        os.makedirs("csv")

    # Check if the Excel file exists
    if not os.path.exists(args.filename):
        print(f"Error: File '{args.filename}' not found")
        sys.exit(1)

    if args.by_team:
        export_by_team(args.filename)
    else:
        export_by_workbook(args.filename)
        
    # Print total player count after processing all workbooks
    print(f'\nTotal players: {player_count}')
    
if __name__ == '__main__':
    main()