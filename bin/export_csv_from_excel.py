#!/usr/bin/env python3.9
# -*- coding: utf-8 -*-
#
# File: export_csv_from_excel.py
# Author: Scott Stonefield
#
# Read in an Excel file and export each sheet as a CSV file
#
import csv
import openpyxl
import os
import argparse
import sys

# Initial player count
player_count = 0

# Header constants needs to match the headers in the Excel file
HEADER_FIRSTNAME = 'Players First Name'
HEADER_LASTNAME = 'Players Last Name'
HEADER_TEAM = 'Players Team'
HEADER_SWEATER = 'Jersey Number'

# Create mapping for header names to original header names
header_mapping = {
    HEADER_FIRSTNAME: 'Firstname',
    HEADER_LASTNAME: 'Lastname',
    HEADER_TEAM: 'Team',
    HEADER_SWEATER: 'Sweater'
}

def clean_name(name, is_last_name=False):
    """Clean a name by removing leading/trailing spaces and capitalizing the first letter."""
    if name is None:
        return None
    name = name.strip()
    if len(name) == 2 and is_last_name:  # if it's a last name and has only two letters
        return name[0].upper() + name[1].lower()
    
    # if name is not a last_namd capitalize both letters unless it matches 'Ty'
    if len(name) == 2 and name.lower() != 'ty':
        return name[0].upper() + name[1].upper()
    
    # if the name has a hyphen, capitalize the first letter after the hyphen
    if '-' in name:
        return '-'.join(word.capitalize() for word in name.split('-'))

    # if the last_name has two c next two each other in the name, capitalize the second c
    if 'mcc' in name.lower() or 'macc' in name.lower():
        print(f"Alert: Found 'cc' in name '{name}'")
        name = name.replace('cc', 'cC')
    # if the name has two words, capitalize each word
    if len(name.split()) == 2:
        return ' '.join(word.capitalize() if word.lower() not in ['iii', 'iv'] else word.upper() for word in name.split())
    return name.capitalize()

def clean_sweater_number(sweater_number):
    """Clean a sweater number by removing non-digit characters."""
    if not sweater_number:
        return '00'
    cleaned_number = ''.join(filter(str.isdigit, str(sweater_number)))
    return cleaned_number if cleaned_number else '00'

def clean_row(row, headers):
    global player_count

    firstname_index = headers.index(HEADER_FIRSTNAME)
    lastname_index = headers.index(HEADER_LASTNAME)
    sweater_index = headers.index(HEADER_SWEATER)

    row = list(row)

    row[firstname_index] = clean_name(row[firstname_index])
    if row[firstname_index] is None:
        print(f"Error: {HEADER_FIRSTNAME} is empty in row '{row[lastname_index]}' -- team: {row[headers.index(HEADER_TEAM)]}")

    # if the last name is all caps then call clean_name with is_last_name=True
    if row[lastname_index] is not None and row[lastname_index].isupper():
        row[lastname_index] = clean_name(row[lastname_index], is_last_name=True)
    # if last name is all lower case, capitalize the first letter
    elif row[lastname_index] is not None and row[lastname_index].islower():
        row[lastname_index] = clean_name(row[lastname_index])
        
    if row[lastname_index] is None:
        print(f"Error: {HEADER_LASTNAME} is empty in row '{row[firstname_index]}' -- team: {row[headers.index(HEADER_TEAM)]}")

    row[sweater_index] = clean_sweater_number(row[sweater_index])

    player_count += 1
    return row

def export_by_workbook(filename):
    global player_count
    
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)
    
    player_count = 0
    # Iterate through each sheet in the workbook
    for sheet in wb:
        # Get the first row (which should contain the headers)
        headers = [cell.value for cell in next(sheet.rows)]

        # Check if the required headers are present
        if not all(header in headers for header in header_mapping.keys()):
            print(f"Error: Required headers not found in sheet '{sheet.title}'")
            continue
        
        # If sheet.title has a space in the name, replace it with an underscore
        sheet.title = sheet.title.replace(' ', '_')
        
        # replace any characters that are not allowed in a filename with an underscore
        sheet.title = ''.join(c if c.isalnum() else '_' for c in sheet.title)
        
        # Create a CSV file for the sheet
        with open(f'csv/{sheet.title}.csv', 'w', newline='') as csv_file:
            # Create a CSV writer
            writer = csv.writer(csv_file)
            # Write the mapped headers to the CSV file
            writer.writerow([header_mapping[header] for header in headers])
            # Iterate through the rest of the rows in the sheet and write them to the CSV file
            for row in sheet.iter_rows(min_row=2, values_only=True):  # start from the second row to skip the header
                cleaned_row = clean_row(row, headers)
                writer.writerow(cleaned_row)

def get_headers(sheet):
    """Get the headers from the first row of the sheet."""
    return [cell.value for cell in next(sheet.rows)]

def create_csv_writer(team_name, headers):
    """Create a new CSV writer for a team."""
    # replace any blank spaces from the team name with an underscore
    file_name = team_name.replace(' ', '_') if team_name else ""
    # replace any characters that are not allowed in a filename with an underscore
    file_name = ''.join(c if c.isalnum() else '_' for c in file_name)
    
    # create a new CSV file and writer
    csv_file = open(f'csv/{file_name}.csv', 'w', newline='')
    writer = csv.writer(csv_file)
    writer.writerow([header_mapping[header] for header in headers])  # write the mapped headers to the CSV file

    return csv_file, writer

def export_by_team(filename):
    global player_count
    
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    player_count = 0
    # Iterate through each sheet in the workbook
    for sheet in wb:
        headers = get_headers(sheet)

        # Check if the required headers are present
        if not all(header in headers for header in header_mapping.keys()):
            print(f"Error: Required headers not found in sheet '{sheet.title}'")
            continue

        # Adjust column indices based on your header structure
        team_index = headers.index(HEADER_TEAM)
        
        last_team_name = None
        csv_file = None
        writer = None
        team_player_count = 0

        # Iterate through the rest of the rows in the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):  # start from the second row to skip the header
            # if the row is empty, skip it
            if row[team_index] is None:
                continue

            team_name = row[team_index]
                
            # If team name changes or is the first row, create a new CSV writer
            if team_name != last_team_name:
                if csv_file is not None:  # if not the first row
                    csv_file.close()  # close the last CSV file
                    if team_player_count < 10:  # if the last team had less than 10 players
                        print(f"Alert: Team {last_team_name} has less than 10 players!")

                print(f"Team:", team_name)
                
                csv_file, writer = create_csv_writer(team_name, headers)

                last_team_name = team_name
                
            cleaned_row = clean_row(row, headers)
            writer.writerow(cleaned_row)  # write the row to the CSV file
            team_player_count += 1 

        if csv_file is not None:
            csv_file.close()  # close the last CSV file
            if team_player_count < 10:  # if the last team had less than 10 players
                print(f"Alert: Team {last_team_name} has less than 10 players!")

def main():
    global player_count
    
    # Command line argument parsing
    parser = argparse.ArgumentParser(description='Export Excel file to CSV.')
    parser.add_argument('filename', help='Name of the Excel file to export.')
    parser.add_argument('--by-team', action='store_true', help='If set, export by team. Otherwise, export by workbook.')

    args = parser.parse_args()

    # Create the "csv" directory if it doesn't exist
    if not os.path.exists("csv"):
        os.makedirs("csv")

    # Check if the Excel file exists
    if not os.path.exists(args.filename):
        print(f"Error: File '{args.filename}' not found")
        sys.exit(1)

    if args.by_team:
        export_by_team(args.filename)
    else:
        export_by_workbook(args.filename)
        
    # Print total player count after processing all workbooks
    print(f'\nTotal players: {player_count}')
    
if __name__ == '__main__':
    main()
