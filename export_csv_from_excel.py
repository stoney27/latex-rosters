#!/usr/bin/env python3.9
# -*- coding: utf-8 -*-
#
# File: export_csv_from_excel.py
# Author: Scott Stonefield
#
# Read in an Excel file and export each sheet as a CSV file
#
import csv
import openpyxl
import os
import argparse
import sys

# Initial player count
player_count = 0


def clean_row(row, headers):
    global player_count

    firstname_index = headers.index('Firstname')
    lastname_index = headers.index('Lastname')
    sweater_index = headers.index('Sweater')

    row = list(row)

    if row[firstname_index] is None:
        print(f"Error: Firstname is empty in row '{row[lastname_index]}' -- team: {row[headers.index('Team')]}")
    else:
        row[firstname_index] = row[firstname_index].strip()
        # Make sure the first letter of the first name is upper case and the rest are lower case
        row[firstname_index] = row[firstname_index].lower()
        row[firstname_index] = row[firstname_index][0].upper() + row[firstname_index][1:]

    if row[lastname_index] is None:
        print(f"Error: Lastname is empty in row '{row[firstname_index]}' -- team: {row[headers.index('Team')]}")
    else:
        row[lastname_index] = row[lastname_index].strip()
        # Make sure the first letter of the last name is upper case and the rest are lower case
        row[lastname_index] = row[lastname_index].lower()
        row[lastname_index] = row[lastname_index][0].upper() + row[lastname_index][1:]

    # Check and clean Sweater number
    if row[sweater_index] is not None:
        row[sweater_index] = ''.join(filter(str.isdigit, str(row[sweater_index])))
        if not row[sweater_index]:
            print(f"Error: Invalid Sweater number in row '{row[firstname_index]} {row[lastname_index]}' -- team: {row[headers.index('Team')]}")
            row[sweater_index] = '00'
    else:
        print(f"Error: Sweater number is empty in row '{row[firstname_index]} {row[lastname_index]}' -- team: {row[headers.index('Team')]}")
        row[sweater_index] = '00'
    
    player_count += 1
    return row

def export_by_workbook(filename):
    global player_count
    
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)
    
    player_count = 0
    # Iterate through each sheet in the workbook
    for sheet in wb:
        # Get the first row (which should contain the headers)
        headers = [cell.value for cell in next(sheet.rows)]

        # Check if the required headers are present
        if not all(header in headers for header in ['Firstname', 'Lastname', 'Team', 'Sweater']):
            print(f"Error: Required headers not found in sheet '{sheet.title}'")
            continue
        
        # If sheet.title has a space in the name, replace it with an underscore
        sheet.title = sheet.title.replace(' ', '_')
        
        # replace any characters that are not allowed in a filename with an underscore
        sheet.title = ''.join(c if c.isalnum() else '_' for c in sheet.title)
        
        # Create a CSV file for the sheet
        with open(f'csv/{sheet.title}.csv', 'w', newline='') as csv_file:
            # Create a CSV writer
            writer = csv.writer(csv_file)
            # Write the headers to the CSV file
            writer.writerow(headers)
            # Iterate through the rest of the rows in the sheet and write them to the CSV file
            for row in sheet.iter_rows(min_row=2, values_only=True):  # start from the second row to skip the header
                writer.writerow(row)

def export_by_team(filename):
    global player_count
    
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    player_count = 0
    # Iterate through each sheet in the workbook
    for sheet in wb:
        # Get the first row (which should contain the headers)
        headers = [cell.value for cell in next(sheet.rows)]

        # Check if the required headers are present
        if not all(header in headers for header in ['Firstname', 'Lastname', 'Team', 'Sweater']):
            print(f"Error: Required headers not found in sheet '{sheet.title}'")
            continue

        # Adjust column indices based on your header structure
        team_index = headers.index('Team')
        firstname_index = headers.index('Firstname')
        lastname_index = headers.index('Lastname')
        
        last_team_name = None
        csv_file = None
        writer = None

        # Iterate through the rest of the rows in the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):  # start from the second row to skip the header
            # if the row is empty, skip it
            if row[team_index] is None:
                continue

            team_name = row[team_index]
                
            # If team name changes or is the first row, create a new CSV writer
            if team_name != last_team_name:
                if csv_file is not None:  # if not the first row
                    csv_file.close()  # close the last CSV file

                # replace any blank spaces from the team name with an underscore
                if team_name is not None:
                    file_name = team_name.replace(' ', '_')
                    # replace any characters that are not allowed in a filename with an underscore
                    file_name = ''.join(c if c.isalnum() else '_' for c in file_name)
                
                # Print the name of the new team
                print(f"Team:", team_name)
                
                # create a new CSV file and writer
                csv_file = open(f'csv/{file_name}.csv', 'w', newline='')
                writer = csv.writer(csv_file)
                writer.writerow(headers)  # write the headers to the CSV file

                last_team_name = team_name
                
            cleaned_row = clean_row(row, headers)
            writer.writerow(cleaned_row)  # write the row to the CSV file

        if csv_file is not None:
            csv_file.close()  # close the last CSV file

def main():
    global player_count
    
    # Command line argument parsing
    parser = argparse.ArgumentParser(description='Export Excel file to CSV.')
    parser.add_argument('filename', help='Name of the Excel file to export.')
    parser.add_argument('--by-team', action='store_true', help='If set, export by team. Otherwise, export by workbook.')

    args = parser.parse_args()

    # Create the "csv" directory if it doesn't exist
    if not os.path.exists("csv"):
        os.makedirs("csv")

    # Check if the Excel file exists
    if not os.path.exists(args.filename):
        print(f"Error: File '{args.filename}' not found")
        sys.exit(1)

    if args.by_team:
        export_by_team(args.filename)
    else:
        export_by_workbook(args.filename)
        
    # Print total player count after processing all workbooks
    print(f'\nTotal players: {player_count}')
    
if __name__ == '__main__':
    main()

