#!/usr/bin/env python3.9
# -*- coding: utf-8 -*-
"""
Created on Sun Jan  8 13:31:20 EST 2023
This script will read an excel file and create a CSV file for each team in the excel file.
The CSV file will contain the team name, the team number, the team's match number, the team's score, and the team's rank.
The CSV file will be named after the team name.
The CSV file will be saved in the same directory called csv_files.

"""
import csv
import openpyxl
import os
import sys
import re

# Check if the required headers are present in the sheet
def check_sheet_titles(sheet):
    # Get the first row (which should contain the headers)
    headers = [cell.value for cell in next(sheet.rows)]
    
    # print headers
    print(f"Headers: {headers}")
    
    # If the number of columes is 3, then check for Name, Team, and Sweater
    if len(headers) == 3 or (len(headers) == 4 and headers[3] == None):
        if not all(header in headers for header in ['Name', 'Team', 'Sweater']):
            print(f"Error: Required headers not found in sheet '{sheet.title}'")
            return False
        else:
            return True

    # Check if the required headers are present
    if not all(header in headers for header in ['Firstname', 'Lastname', 'Team', 'Sweater']):
        # Check for lowercase headers
        for header in headers:
            if header.islower():
                headers[headers.index(header)] = header.capitalize()

        # Check for alternative variations of the headers
        if 'First Name' in headers:
            headers[headers.index('First Name')] = 'Firstname'
        if 'Last Name' in headers:
            headers[headers.index('Last Name')] = 'Lastname'

        if not all(header in headers for header in ['Firstname', 'Lastname', 'Team', 'Sweater']):
            print(f"Error: Required headers not found in sheet '{sheet.title}'")
            return False
    return True

# Function to seperate first and last name
def separate_first_last_name(name):

    # Remove the string ' - Goalie' from the Full Name if it exists
    name = name.replace(' - Goalie', '')

    # Check for the number of spaces in the name
    splitname = name.split()
    case = len(splitname)
    if case == 1:
        # Only one name, return it
        return splitname[0].capitalize(), ''
    if case == 2:
        return splitname[0].capitalize(), splitname[1].capitalize()
    if case == 3:
        if splitname[2] == 'III' or splitname[2] == 'II' or splitname[2] == 'IV' or splitname[1] == 'Van' or splitname[2] == 'Jr.':
            return splitname[0].capitalize(), splitname[1].capitalize() + ' ' + splitname[2]
        if splitname[1].find('(') != -1:
            return splitname[0].capitalize() + ' ' + splitname[1], splitname[2].capitalize()
        else:
            return splitname[0].capitalize() + ' ' + splitname[1].capitalize(), splitname[2].capitalize()
    else:
        print("Error: %s" % name)
        return '', ''

# Function to Fix/Standardize the team names
def fix_team_name(team_name):
    # Set the team name to uppercase
    team_name = team_name.upper()

    # Make the team name has a space after <digit>U or <digit><digit>U
    if re.match(r'\d+U\w+', team_name):
        team_name = team_name.replace('U', 'U ')
    
    # Camel case the team name after the space unless it is AAA
    if team_name.find('AAA') == -1:
        (team_age_group, t_name) = team_name.split(' ', 1)
        team_name = team_age_group + ' ' + t_name.title()

    return team_name


# process the teams and return a dictionary of the rosters
def process_teams(workbook, sheet_name):

    # Get the sheet
    sheet = workbook[sheet_name]

    # Create a dictionary to hold the team rosters
    rosters = {}

    # Iterate through the rows of the sheet
    for row in sheet.iter_rows():

        # Skip the first row (which contains the headers)
        if row[0].row == 1:
            continue
        
        # Skip blank rows
        if row[0].value is None or row[0].value == '':
            continue

        # if row has four columns then it has firstname lastname team and sweater
        if len(row) == 4 and row[3].value is not None:
            # Get the values in the First Name, Last Name, Team, and Sweater Name columns
            first_name, last_name, team, sweater_name = [cell.value for cell in row]
        else:
            # Get the values in the Full Name, Team, and Sweater Name columns
            full_name, team, sweater_name = [cell.value for cell in row[:3]]

            # Split the full name into first and last name
            (first_name, last_name) = separate_first_last_name(full_name)

        # Add the player to the roster for their team
        if team not in rosters:
            rosters[team] = []
        rosters[team].append({
            'Firstname': first_name,
            'Lastname': last_name,
            'Team': team,
            'Sweater': sweater_name
        })

    return rosters

# Write out the rosters to CSV files
def write_rosters_to_csv(rosters):

    # Create the "csv" directory if it doesn't already exist
    if not os.path.exists('csv'):
        os.makedirs('csv')

    # Write out the CSV files for each team
    for team, roster in rosters.items():
        # Print the team name
        print(team)

        # If the team name has a space in it, replace it with an underscore
        if ' ' in team:
            team = team.replace(' ', '_')

        with open(f'csv/{team}.csv', 'w', newline='') as csv_file:
            fieldnames = ['Firstname', 'Lastname', 'Team', 'Sweater']
            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
            writer.writeheader()
            for player in roster:
                writer.writerow(player)

# Main function
def main():
    # Get the file name and sheet name from the command line arguments
    file_name = sys.argv[1]
    try:
        sheet_name = sys.argv[2]
    except IndexError:
        sheet_name = 'Sheet1'
    
    # Check to see if file_name exists if not exit
    if not os.path.exists(file_name):
        print(f"Error: File '{file_name}' does not exist.")
        exit(1)

    # Open the Excel file
    workbook = openpyxl.load_workbook(file_name)

    # Read the number of sheets in the workbook
    number_of_sheets = len(workbook.sheetnames)
    print(f"There are {number_of_sheets} sheets in the workbook.")

    teams = {}
    # If only one sheet then process all teams on same sheet
    if number_of_sheets == 1:
        # Check if the sheet has the required headers
        if not check_sheet_titles(workbook[sheet_name]):
            print(f"Error: Required headers not found in sheet '{sheet_name}'")
            exit(1)

        teams = process_teams(workbook, sheet_name)
    else:
        # Process each sheet in the workbook
        for sheet in workbook:
            # Check if the sheet has the required headers
            if not check_sheet_titles(sheet):
                print(f"Error: Required headers not found in sheet '{sheet.title}'")
                continue
            team = process_teams(workbook, sheet.title)
            # append team to teams
            teams.update(team)

    # Write out teams to their csv file
    write_rosters_to_csv(teams)


if __name__ == "__main__":
    main()
    exit(0)
